#!/usr/bin/env python3
"""
Simula el bucketing + normalización Depicenter con el fix aplicado.
"""
from openpyxl import load_workbook
from datetime import date, timedelta
from collections import defaultdict
import re, unicodedata, sys

PATH = sys.argv[1] if len(sys.argv) > 1 else r"C:\Users\ADMIN\Downloads\ReporteDisparos-2026-06-01.xlsx"
LECTURAS_PATH = r"C:\Users\ADMIN\Downloads\Depicenter_Formato_Oficial_25_30_Mayo_2026.xlsx"

HEADER_SKIP = {"SUCURSAL","OPERADORA","OPERADOR","EQUIPO","EQUIPOID","CABINA","PULSOS","ESTADO","FALLAS","SERIAL","SEMANA","FECHA","CLIENTE","TRATAMIENTO","POTENCIA","SPOT","DISPAROS","SECUENCIAL","CONTACTO","TOTAL","TOTALES"}
BRAND_PREFIX_RE = re.compile(r'^(?:cibao\s+spa\s+l[aá]ser|cibao\s+spa\s+laser|depicenter|csl)\s*[-–—]?\s*', re.I)

def clean_upper(s):
    if s is None: return ""
    s = unicodedata.normalize("NFD", str(s))
    s = "".join(c for c in s if unicodedata.category(c) != "Mn")
    return re.sub(r"\s+", " ", s.upper()).strip()

def normalize_sucursal(value):
    s = "" if value is None else str(value).strip()
    if not s: return ""
    up_raw = clean_upper(s)
    if not up_raw or up_raw in HEADER_SKIP: return ""
    # Depicenter FIRST
    if "DEPICENTER" in up_raw: return "DEPICENTER"
    if "SKIN" in up_raw and "LASER" in up_raw: return "DEPICENTER"
    # Cibao
    stripped = BRAND_PREFIX_RE.sub("", s).strip() or s
    up = clean_upper(stripped)
    if not up or up in HEADER_SKIP: return ""
    if "JARDINES" in up: return "LOS JARDINES"
    if up == "R VIDAL" or any(t in up for t in ("RAFAEL","VIDAL","PLAZA","MEDITERR")): return "RAFAEL VIDAL"
    if ("VILLA" in up and "OLGA" in up) or up == "V OLGA": return "VILLA OLGA"
    if "LA VEGA" in up: return "LA VEGA"
    return up

# ── Lecturas Depicenter
print("=" * 70)
print(f"LECTURAS: {LECTURAS_PATH}")
wb = load_workbook(LECTURAS_PATH, data_only=True)
ws = wb["Equipos"]
matrix = [list(row) for row in ws.iter_rows(values_only=True)]
# Buscar header row
header_row = -1
for i, row in enumerate(matrix[:10]):
    if str(row[0] or "").strip().lower() in ("equipo","equipo_id","equipoid"):
        header_row = i; break
print(f"Header row encontrada (0-idx): {header_row}")
headers = [str(c or "").strip().lower() for c in matrix[header_row]]
print(f"Headers: {headers}")
parsed_equipos = 0
for row in matrix[header_row+1:]:
    if all(c is None or c == "" for c in row): continue
    obj = dict(zip(headers, row))
    eq = str(obj.get("equipo","")).strip()
    if not eq: continue
    suc = obj.get("sucursal","")
    op = obj.get("operadora","")
    print(f"  Equipo {eq:>3} | Sucursal '{suc}' -> '{normalize_sucursal(suc)}' | Operadora '{op}'")
    parsed_equipos += 1
print(f"Total equipos parseados: {parsed_equipos}")

# ── AgendaPro Depicenter
print("\n" + "=" * 70)
print(f"AGENDAPRO: {PATH}")
wb = load_workbook(PATH, data_only=True)
ws = wb["Detalle Disparos tratamientos"]
print(f"Fila 1: {ws.cell(row=1, column=1).value}")
print(f"Fila 4: {[ws.cell(row=4, column=c).value for c in range(1, 11)]}")
print(f"Fila 5 ejemplo: Sucursal='{ws.cell(row=5, column=6).value}' Operadora='{ws.cell(row=5, column=5).value}'")
norm = normalize_sucursal(ws.cell(row=5, column=6).value)
print(f"Sucursal normalizada: '{norm}'")
print("EXPECTED: 'DEPICENTER'")
print("RESULT:", "PASS" if norm == "DEPICENTER" else "FAIL")
