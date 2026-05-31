#!/usr/bin/env python3
"""
Simula el bucketing por semana operativa (lunes-sábado) que hace
Cuadre Semanal cuando se sube un AgendaPro. Predice cuántas filas
deberían terminar en csl_operator_shots.
"""
from openpyxl import load_workbook
from datetime import date, timedelta
from collections import defaultdict
import sys, re, unicodedata

PATH = sys.argv[1] if len(sys.argv) > 1 else r"C:\Users\ADMIN\Downloads\ReporteDisparos-2026-05-24.xlsx"

HEADER_SKIP = {
    "SUCURSAL", "OPERADORA", "OPERADOR", "EQUIPO", "EQUIPOID",
    "CABINA", "PULSOS", "ESTADO", "FALLAS", "SERIAL", "SEMANA",
    "FECHA", "CLIENTE", "TRATAMIENTO", "POTENCIA", "SPOT", "DISPAROS",
    "SECUENCIAL", "CONTACTO", "TOTAL", "TOTALES",
}

BRAND_PREFIX_RE = re.compile(r'^(?:cibao\s+spa\s+l[aá]ser|cibao\s+spa\s+laser|depicenter|csl)\s*[-–—]?\s*', re.I)

OPERADORA_ALIASES = {
    "KATHERINE": "KATHERIN", "EMELY": "EMELI", "RIQUELMI": "ROQUELMI",
    "YESICA": "YESSICA", "SAOMY": "SAHOMY",
}
OPERADORA_SKIP = {
    "SISTEMA", "SYSTEM", "ADMIN",
    "OPERADORA", "OPERADOR", "EQUIPO", "EQUIPOID", "SUCURSAL",
}

def clean_upper(s):
    s = "" if s is None else str(s)
    s = unicodedata.normalize("NFD", s)
    s = "".join(c for c in s if unicodedata.category(c) != "Mn")
    return re.sub(r"\s+", " ", s.upper()).strip()

def normalize_sucursal(value):
    s = ("" if value is None else str(value)).strip()
    if not s: return ""
    stripped = BRAND_PREFIX_RE.sub("", s).strip() or s
    up = clean_upper(stripped)
    if not up or up in HEADER_SKIP: return ""
    if "JARDINES" in up: return "LOS JARDINES"
    if up == "R VIDAL" or any(t in up for t in ("RAFAEL", "VIDAL", "PLAZA", "MEDITERR")):
        return "RAFAEL VIDAL"
    if ("VILLA" in up and "OLGA" in up) or up == "V OLGA": return "VILLA OLGA"
    if "LA VEGA" in up: return "LA VEGA"
    if up == "DEPICENTER": return "LA VEGA"
    return up

def normalize_operadora(value):
    up = clean_upper(value)
    if not up or up in OPERADORA_SKIP: return ""
    return OPERADORA_ALIASES.get(up, up)

def get_operational_week(d):
    if not isinstance(d, date): return None
    dow = d.isoweekday()  # 1=Mon, 7=Sun
    if dow == 7:
        monday = d + timedelta(days=1)
    else:
        monday = d - timedelta(days=dow - 1)
    return (monday.isoformat(), (monday + timedelta(days=5)).isoformat())

def parse_disparos(raw):
    if raw is None: return None
    if isinstance(raw, (int, float)): return int(raw) if raw and raw == raw else None
    s = str(raw).strip()
    if not s: return None
    total = 0
    for part in s.split(","):
        cleaned = re.sub(r"[^\d-]", "", part.strip())
        if not cleaned: return None
        try: total += int(cleaned)
        except ValueError: return None
    return total

def to_iso(val):
    if isinstance(val, date): return val.isoformat()
    s = str(val or "").strip()
    if not s: return None
    m = re.match(r"^(\d{4})-(\d{2})-(\d{2})", s)
    if m: return f"{m.group(1)}-{m.group(2)}-{m.group(3)}"
    m = re.match(r"^(\d{1,2})[/-](\d{1,2})[/-](\d{4})", s)
    if m: return f"{m.group(3)}-{m.group(2).zfill(2)}-{m.group(1).zfill(2)}"
    return None

print(f"Reading {PATH}\n")
wb = load_workbook(PATH, read_only=True, data_only=True)
sheet_names = wb.sheetnames
target = next((n for n in sheet_names if "detalle" in n.lower() and "disparos" in n.lower()), sheet_names[0])
print(f"Hoja: {target}")
ws = wb[target]
rows = list(ws.iter_rows(values_only=True))
print(f"Total filas leídas: {len(rows)}")
print(f'Fila 1 col A: "{rows[0][0] if rows else ""}"')

# Detect header row
header_row = -1
for i, row in enumerate(rows):
    first = "" if not row or row[0] is None else str(row[0]).lower()
    if "secuencial" in first:
        header_row = i
        break
print(f"Header row (0-idx): {header_row}")

valid = []
total_with_seq = 0
skipped = defaultdict(int)
for i in range(header_row + 1, len(rows)):
    row = rows[i]
    if not row: continue
    seq_raw = "" if row[0] is None else str(row[0]).strip()
    if not seq_raw: continue
    try:
        int(seq_raw.replace(",", ""))
    except ValueError:
        skipped["no_secuencial"] += 1
        continue
    total_with_seq += 1
    operadora = ("" if row[4] is None else str(row[4])).strip()
    sucursal = ("" if row[5] is None else str(row[5])).strip()
    fecha_iso = to_iso(row[9])
    disparos = parse_disparos(row[8])
    if not fecha_iso: skipped["sin_fecha"] += 1; continue
    suc_norm = normalize_sucursal(sucursal)
    op_norm = normalize_operadora(operadora)
    if not suc_norm: skipped["sin_sucursal"] += 1; continue
    if not op_norm: skipped["sin_operadora"] += 1; continue
    if disparos is None or disparos <= 0: skipped["sin_disparos"] += 1; continue
    valid.append({
        "fecha": fecha_iso, "sucursal": sucursal, "suc_norm": suc_norm,
        "operadora": operadora, "op_norm": op_norm, "disparos": disparos,
    })

print(f"\nFilas con secuencial numérico: {total_with_seq}")
print(f"Filas válidas tras normalización: {len(valid)}")
print(f"Filtradas: {dict(skipped)}")

# Bucketing
buckets = defaultdict(lambda: {"by_key": defaultdict(lambda: {"sesiones": 0, "disparos": 0, "suc_norm": "", "op_norm": ""}), "sesiones": 0, "disparos": 0, "period_end": None})
for r in valid:
    y, m, d = (int(x) for x in r["fecha"].split("-"))
    week = get_operational_week(date(y, m, d))
    if not week: continue
    ps, pe = week
    b = buckets[ps]
    b["period_end"] = pe
    key = f"{r['suc_norm']}|{r['op_norm']}"
    cell = b["by_key"][key]
    cell["sesiones"] += 1
    cell["disparos"] += r["disparos"]
    cell["suc_norm"] = r["suc_norm"]
    cell["op_norm"] = r["op_norm"]
    b["sesiones"] += 1
    b["disparos"] += r["disparos"]

print(f"\n=== {len(buckets)} SEMANAS OPERATIVAS DETECTADAS ===\n")
total_shots = 0
total_sesiones = 0
total_disparos = 0
for ps in sorted(buckets.keys()):
    b = buckets[ps]
    print(f"Semana {ps} -> {b['period_end']}")
    print(f"  Sesiones: {b['sesiones']}  |  Disparos: {b['disparos']:,}  |  Filas operator_shots: {len(b['by_key'])}")
    for key in sorted(b["by_key"].keys()):
        cell = b["by_key"][key]
        print(f"    | {cell['suc_norm']:<15} {cell['op_norm']:<12} {cell['sesiones']:>4} ses  {cell['disparos']:>10,} disp")
    total_shots += len(b["by_key"])
    total_sesiones += b["sesiones"]
    total_disparos += b["disparos"]

print(f"\n=== RESUMEN GLOBAL ===")
print(f"Semanas operativas detectadas: {len(buckets)}")
print(f"Filas que se insertarán en csl_operator_shots: {total_shots}")
print(f"Total sesiones: {total_sesiones}")
print(f"Total disparos: {total_disparos:,}")
