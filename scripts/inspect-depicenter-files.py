#!/usr/bin/env python3
"""
Inspecciona los 4 Excel de Depicenter que el usuario reporta no funcionar.
Compara estructura contra el formato Cibao para detectar diferencias.
"""
from openpyxl import load_workbook
from pathlib import Path

FILES = [
    r"C:\Users\ADMIN\Downloads\Depicenter_Formato_Oficial_11_16_Mayo_2026.xlsx",
    r"C:\Users\ADMIN\Downloads\Depicenter_Formato_Oficial_18_23_Mayo_2026.xlsx",
    r"C:\Users\ADMIN\Downloads\Depicenter_Formato_Oficial_25_30_Mayo_2026.xlsx",
    r"C:\Users\ADMIN\Downloads\ReporteDisparos-2026-06-01.xlsx",
]

for f in FILES:
    p = Path(f)
    print("=" * 80)
    print(f"ARCHIVO: {p.name}")
    print(f"Tamano: {p.stat().st_size:,} bytes")
    try:
        wb = load_workbook(p, read_only=False, data_only=True)
        print(f"Hojas: {wb.sheetnames}")
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            print(f"\n  Hoja '{sheet_name}':")
            try:
                print(f"    Max row: {ws.max_row}, Max col: {ws.max_column}")
            except Exception:
                pass
            # Primeras 6 filas
            rows = []
            for i, row in enumerate(ws.iter_rows(values_only=True)):
                if i >= 6: break
                rows.append(row)
            for i, row in enumerate(rows):
                vals = [str(c)[:30] if c is not None else "" for c in row]
                print(f"    Fila {i+1}: {vals}")
    except Exception as e:
        print(f"ERROR: {e}")
    print()
