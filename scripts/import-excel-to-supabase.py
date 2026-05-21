import argparse
import json
import os
import re
import sys
import time
import urllib.error
import urllib.parse
import urllib.request
from datetime import date, datetime
from pathlib import Path
from typing import Any

from openpyxl import load_workbook


ROOT = Path(__file__).resolve().parents[1]
DEFAULT_EXCEL_PATH = ROOT / "tmp_sistema_csl_import.xlsx"


TABLE_KEYS = {
    "csl_sucursales": "codigo",
    "csl_equipos": "equipo_id",
    "csl_tecnicos": "codigo",
    "csl_piezas": "pieza",
    "csl_reportes": "report_id",
    "csl_inventario": "item_id",
    "csl_operadoras": "operadora_id",
    "csl_lecturas_semanales": "lectura_id",
    "csl_sesiones_cliente": "sesion_id",
    "csl_auditorias_semanales": "auditoria_id",
    "csl_credenciales": "credencial_id",
    "csl_solicitudes_empleo": "solicitud_id",
    "csl_empleados": "empleado_id",
}


def load_env() -> dict[str, str]:
    env = dict(os.environ)
    env_path = ROOT / ".env.local"
    if env_path.exists():
        for line in env_path.read_text(encoding="utf-8").splitlines():
            line = line.strip()
            if not line or line.startswith("#") or "=" not in line:
                continue
            key, value = line.split("=", 1)
            env.setdefault(key.strip(), value.strip().strip('"').strip("'"))
    return env


def clean_text(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value).strip()


def clean_number(value: Any, fallback: float = 0) -> float:
    if value is None or value == "":
        return fallback
    try:
        return float(value)
    except (TypeError, ValueError):
        return fallback


def clean_date(value: Any) -> str | None:
    if value is None or value == "":
        return None
    if isinstance(value, datetime):
        return value.date().isoformat()
    if isinstance(value, date):
        return value.isoformat()
    text = clean_text(value)
    match = re.search(r"\d{4}-\d{2}-\d{2}", text)
    return match.group(0) if match else None


def parse_json(value: Any, fallback: Any) -> Any:
    text = clean_text(value)
    if not text:
        return fallback
    try:
        return json.loads(text)
    except json.JSONDecodeError:
        return fallback


def row_dict(headers: list[str], row: tuple[Any, ...]) -> dict[str, Any]:
    return {headers[index]: row[index] if index < len(row) else None for index in range(len(headers))}


def read_sheet(ws) -> list[dict[str, Any]]:
    rows = list(ws.iter_rows(values_only=True))
    if len(rows) < 2:
        return []
    headers = [clean_text(header) for header in rows[0]]
    return [row_dict(headers, row) for row in rows[1:] if any(cell not in (None, "") for cell in row)]


def map_sucursales(records: list[dict[str, Any]]) -> list[dict[str, Any]]:
    return [
        {
            "codigo": clean_text(record.get("Codigo")),
            "nombre": clean_text(record.get("Nombre")),
            "ciudad": clean_text(record.get("Ciudad")),
            "direccion": clean_text(record.get("Direccion")),
            "estado": clean_text(record.get("Estado")) or "Activa",
            "notas": clean_text(record.get("Notas")),
            "correo": clean_text(record.get("Correo")),
        }
        for record in records
        if clean_text(record.get("Codigo"))
    ]


def map_equipos(records: list[dict[str, Any]]) -> list[dict[str, Any]]:
    return [
        {
            "equipo_id": clean_text(record.get("EquipoID")),
            "sucursal": clean_text(record.get("Sucursal")),
            "empresa": clean_text(record.get("Empresa")),
            "domicilio": clean_text(record.get("Domicilio")),
            "modelo": clean_text(record.get("Modelo")),
            "serie": clean_text(record.get("Serie")),
            "numero": clean_text(record.get("Numero")),
            "p_cabeza": clean_number(record.get("P_Cabeza")),
            "p_totales": clean_number(record.get("P_Totales")),
            "max_cabeza": clean_number(record.get("Max_Cabeza"), 6000000),
            "estado": clean_text(record.get("Estado")) or "Activo",
            "observaciones": clean_text(record.get("Observaciones")),
        }
        for record in records
        if clean_text(record.get("EquipoID"))
    ]


def map_tecnicos(records: list[dict[str, Any]]) -> list[dict[str, Any]]:
    return [
        {
            "codigo": clean_text(record.get("Codigo")),
            "nombre": clean_text(record.get("Nombre")),
            "telefono": clean_text(record.get("Telefono")),
            "correo": clean_text(record.get("Correo")),
            "estado": clean_text(record.get("Estado")) or "Activo",
            "notas": clean_text(record.get("Notas")),
        }
        for record in records
        if clean_text(record.get("Codigo"))
    ]


def map_piezas(records: list[dict[str, Any]]) -> list[dict[str, Any]]:
    return [
        {
            "pieza": clean_text(record.get("Pieza")),
            "categoria": clean_text(record.get("Categoria")),
            "prioridad": clean_text(record.get("Prioridad")) or "Media",
            "tipo": clean_text(record.get("Tipo")) or "Consumible",
            "funcion": clean_text(record.get("Funcion")),
            "fallas_comunes": clean_text(record.get("FallasComunes")),
            "activa": clean_text(record.get("Activa")) or "Sí",
        }
        for record in records
        if clean_text(record.get("Pieza"))
    ]


def map_reportes(records: list[dict[str, Any]]) -> list[dict[str, Any]]:
    return [
        {
            "report_id": clean_text(record.get("ID")),
            "fecha": clean_date(record.get("Fecha")),
            "equipo_id": clean_text(record.get("EquipoID")),
            "sucursal": clean_text(record.get("Sucursal")),
            "empresa": clean_text(record.get("Empresa")),
            "cliente": clean_text(record.get("Cliente")),
            "domicilio": clean_text(record.get("Domicilio")),
            "ciudad": clean_text(record.get("Ciudad")),
            "modelo": clean_text(record.get("Modelo")),
            "serie": clean_text(record.get("Serie")),
            "numero": clean_text(record.get("Numero")),
            "tipo": clean_text(record.get("Tipo")) or "Preventivo",
            "estado_equipo": clean_text(record.get("EstadoEquipo")) or "Operativo",
            "prioridad": clean_text(record.get("Prioridad")) or "Baja",
            "problema": clean_text(record.get("Problema")),
            "correccion": clean_text(record.get("Correccion")),
            "observaciones": clean_text(record.get("Observaciones")),
            "checklist": clean_text(record.get("Checklist")),
            "p_cabeza": clean_number(record.get("P_Cabeza")),
            "p_totales": clean_number(record.get("P_Totales")),
            "atendio": clean_text(record.get("Atendio")),
            "piezas_json": clean_text(record.get("PiezasJSON")) or "[]",
            "firma_cliente": clean_text(record.get("FirmaCliente")),
            "firma_tecnico": clean_text(record.get("FirmaTecnico")),
            "fotos": clean_text(record.get("Fotos")),
        }
        for record in records
        if clean_text(record.get("ID"))
    ]


def map_inventario(records: list[dict[str, Any]]) -> list[dict[str, Any]]:
    return [
        {
            "item_id": clean_text(record.get("ItemID")),
            "codigo_barras": clean_text(record.get("CodigoBarras")),
            "pieza": clean_text(record.get("Pieza")),
            "categoria": clean_text(record.get("Categoria")),
            "marca": clean_text(record.get("Marca")),
            "modelo": clean_text(record.get("Modelo")),
            "numero_parte": clean_text(record.get("NumeroParte")),
            "precio_compra": clean_number(record.get("PrecioCompra")),
            "precio_compra_mercado": clean_number(record.get("PrecioCompraMercado")),
            "precio_venta": clean_number(record.get("PrecioVenta")),
            "stock_rafael_vidal": clean_number(record.get("StockRafaelVidal")),
            "stock_los_jardines": clean_number(record.get("StockLosJardines")),
            "stock_villa_olga": clean_number(record.get("StockVillaOlga")),
            "stock_la_vega": clean_number(record.get("StockLaVega")),
            "stock_minimo": clean_number(record.get("StockMinimo")),
            "proveedor": clean_text(record.get("Proveedor")),
            "estado": clean_text(record.get("Estado")) or "Activo",
            "observaciones": clean_text(record.get("Observaciones")),
        }
        for record in records
        if clean_text(record.get("ItemID"))
    ]


def map_operadoras(records: list[dict[str, Any]]) -> list[dict[str, Any]]:
    return [
        {
            "operadora_id": clean_text(record.get("OperadoraID")),
            "nombre": clean_text(record.get("Nombre")),
            "sucursal": clean_text(record.get("Sucursal")),
            "estado": clean_text(record.get("Estado")) or "Activa",
            "notas": clean_text(record.get("Notas")),
        }
        for record in records
        if clean_text(record.get("OperadoraID"))
    ]


def map_lecturas(records: list[dict[str, Any]]) -> list[dict[str, Any]]:
    return [
        {
            "lectura_id": clean_text(record.get("LecturaID")),
            "fecha_semana": clean_date(record.get("FechaSemana")),
            "equipo_id": clean_text(record.get("EquipoID")),
            "sucursal": clean_text(record.get("Sucursal")),
            "cabina": clean_text(record.get("Cabina")),
            "operadora_id": clean_text(record.get("OperadoraID")),
            "lectura_inicial": clean_number(record.get("LecturaInicial")),
            "lectura_final": clean_number(record.get("LecturaFinal")),
            "diferencia_real": clean_number(record.get("DiferenciaReal")),
            "observaciones": clean_text(record.get("Observaciones")),
        }
        for record in records
        if clean_text(record.get("LecturaID"))
    ]


def map_sesiones(ws) -> list[dict[str, Any]]:
    rows = list(ws.iter_rows(values_only=True))[1:]
    mapped = []
    for index, row in enumerate(rows, start=1):
        if not any(cell not in (None, "") for cell in row):
            continue
        first = clean_text(row[0] if len(row) > 0 else None)
        second = clean_text(row[1] if len(row) > 1 else None)
        if clean_date(first) and second:
            fecha = clean_date(first)
            equipo_id = second
            sucursal = row[2] if len(row) > 2 else None
            cabina = row[3] if len(row) > 3 else None
            operadora = row[4] if len(row) > 4 else None
            cliente = row[5] if len(row) > 5 else None
            area = row[6] if len(row) > 6 else None
            disparos = row[7] if len(row) > 7 else None
            duracion = row[8] if len(row) > 8 else None
            observaciones = row[10] if len(row) > 10 else None
            sesion_id = f"ses_{fecha.replace('-', '')}_{index:05d}"
        else:
            sesion_id = clean_text(row[0] if len(row) > 0 else None) or f"ses_{index:05d}"
            fecha = clean_date(row[1] if len(row) > 1 else None)
            sucursal = row[2] if len(row) > 2 else None
            cabina = row[3] if len(row) > 3 else None
            operadora = row[4] if len(row) > 4 else None
            cliente = row[5] if len(row) > 5 else None
            area = row[6] if len(row) > 6 else None
            disparos = row[7] if len(row) > 7 else None
            duracion = row[8] if len(row) > 8 else None
            equipo_id = row[9] if len(row) > 9 else None
            observaciones = row[10] if len(row) > 10 else None
        mapped.append(
            {
                "sesion_id": sesion_id,
                "fecha": fecha,
                "sucursal": clean_text(sucursal),
                "cabina": clean_text(cabina),
                "operadora_id": clean_text(operadora),
                "cliente": clean_text(cliente),
                "area_trabajada": clean_text(area),
                "disparos_reportados": clean_number(disparos),
                "duracion": clean_number(duracion) if duracion not in (None, "") else None,
                "equipo_id": clean_text(equipo_id),
                "observaciones": clean_text(observaciones),
            }
        )
    return mapped


def build_auditorias(lecturas: list[dict[str, Any]], sesiones: list[dict[str, Any]]) -> list[dict[str, Any]]:
    auditorias = []
    for lectura in lecturas:
        fecha_semana = clean_date(lectura.get("fecha_semana"))
        if not fecha_semana:
            continue
        semana_inicio = datetime.fromisoformat(fecha_semana).date()
        semana_fin = semana_inicio.fromordinal(semana_inicio.toordinal() + 6)
        operadora = clean_text(lectura.get("operadora_id")).lower()
        sesiones_semana = []
        for sesion in sesiones:
            fecha_sesion = clean_date(sesion.get("fecha"))
            if not fecha_sesion:
                continue
            fecha = datetime.fromisoformat(fecha_sesion).date()
            misma_operadora = clean_text(sesion.get("operadora_id")).lower() == operadora
            if misma_operadora and semana_inicio <= fecha <= semana_fin:
                sesiones_semana.append(sesion)

        pulsos_reales = clean_number(lectura.get("diferencia_real"))
        pulsos_reportados = sum(clean_number(sesion.get("disparos_reportados")) for sesion in sesiones_semana)
        diferencia = pulsos_reportados - pulsos_reales
        porcentaje = round((diferencia / pulsos_reales) * 100, 2) if pulsos_reales else 0
        desviacion_abs = abs(porcentaje)
        alerta = "OK" if desviacion_abs <= 5 else "Advertencia" if desviacion_abs <= 15 else "Critico"
        equipo_id = clean_text(lectura.get("equipo_id"))
        auditorias.append(
            {
                "auditoria_id": f"aud_{fecha_semana.replace('-', '')}_{equipo_id or clean_text(lectura.get('lectura_id'))}",
                "fecha_semana": fecha_semana,
                "equipo_id": equipo_id,
                "sucursal": clean_text(lectura.get("sucursal")),
                "pulsos_reales": pulsos_reales,
                "pulsos_reportados": pulsos_reportados,
                "diferencia": diferencia,
                "porcentaje_desviacion": porcentaje,
                "alerta": alerta,
                "observaciones": f"Generado desde lecturas y sesiones: {len(sesiones_semana)} sesiones",
            }
        )
    return auditorias


def map_credenciales(records: list[dict[str, Any]]) -> list[dict[str, Any]]:
    return [
        {
            "credencial_id": clean_text(record.get("CredencialID")),
            "sucursal": clean_text(record.get("Sucursal")),
            "area": clean_text(record.get("Area")),
            "equipo": clean_text(record.get("Equipo")),
            "sistema": clean_text(record.get("Sistema")),
            "usuario": clean_text(record.get("Usuario")),
            "contrasena": clean_text(record.get("Contrasena")),
            "pin": clean_text(record.get("PIN")),
            "url": clean_text(record.get("URL")),
            "correo": clean_text(record.get("Correo")),
        }
        for record in records
        if clean_text(record.get("CredencialID"))
    ]


def map_solicitudes(records: list[dict[str, Any]]) -> tuple[list[dict[str, Any]], list[dict[str, Any]]]:
    solicitudes = []
    empleados = []
    for record in records:
        solicitud_id = clean_text(record.get("SolicitudID"))
        if not solicitud_id:
            continue
        payload = parse_json(record.get("Experiencia"), {})
        if not payload:
            payload = parse_json(record.get("Observaciones"), {})
        solicitud = {
            "solicitud_id": solicitud_id,
            "fecha_solicitud": clean_date(record.get("FechaSolicitud")),
            "estado": clean_text(record.get("Estado")) or "Pendiente",
            "puesto_solicitado": clean_text(record.get("PuestoSolicitado")) or clean_text(payload.get("puestoSolicitado")),
            "nombre": clean_text(record.get("Nombre")) or clean_text(payload.get("nombre")),
            "apellido": clean_text(record.get("Apellido")) or clean_text(payload.get("apellido")),
            "cedula": clean_text(record.get("Cedula")) or clean_text(payload.get("cedula")),
            "email": clean_text(record.get("Email")) or clean_text(payload.get("email")),
            "telefono": clean_text(record.get("Telefono")) or clean_text(payload.get("celular") or payload.get("telefonoResidencia")),
            "fecha_nacimiento": clean_date(record.get("FechaNacimiento")) or clean_date(payload.get("fechaNacimiento")),
            "sexo": clean_text(record.get("Sexo")) or clean_text(payload.get("sexo")),
            "nacionalidad": clean_text(record.get("Nacionalidad")) or clean_text(payload.get("nacionalidad")),
            "provincia": clean_text(record.get("Provincia")) or clean_text(payload.get("provincia")),
            "ciudad": clean_text(record.get("Ciudad")) or clean_text(payload.get("ciudad")),
            "direccion": clean_text(record.get("Direccion")) or clean_text(payload.get("direccion")),
            "experiencia": clean_text(payload.get("experiencia")) if isinstance(payload.get("experiencia"), str) else "",
            "salario": clean_number(record.get("Salario") or payload.get("pretensionesSalariales")),
            "nivel_educacion": clean_text(record.get("NivelEducacion")) or clean_text(payload.get("nivelEducacion")),
            "especialidad": clean_text(record.get("Especialidad")) or clean_text(payload.get("especialidad")),
            "documentos_adjuntos": parse_json(record.get("DocumentosAdjuntos"), []),
            "firma_digital": clean_text(record.get("FirmaDigital")) or clean_text(payload.get("firma")),
            "observaciones": clean_text(record.get("Observaciones")) if not clean_text(record.get("Observaciones")).startswith("{") else clean_text(payload.get("observaciones")),
            "fecha_revision": clean_date(record.get("FechaRevision")),
            "revisado_por": clean_text(record.get("RevisadoPor")),
            "payload_json": payload or {"id": solicitud_id},
        }
        solicitudes.append(solicitud)
        if solicitud["estado"] == "Aprobado":
            empleado = dict(solicitud)
            empleado["empleado_id"] = solicitud_id
            empleados.append(empleado)
    return solicitudes, empleados


def build_payloads(excel_path: Path) -> dict[str, list[dict[str, Any]]]:
    workbook = load_workbook(excel_path, data_only=True)
    payloads: dict[str, list[dict[str, Any]]] = {}
    sheet_map = {
        "Sucursales": ("csl_sucursales", lambda ws: map_sucursales(read_sheet(ws))),
        "Equipos": ("csl_equipos", lambda ws: map_equipos(read_sheet(ws))),
        "Tecnicos": ("csl_tecnicos", lambda ws: map_tecnicos(read_sheet(ws))),
        "Catalogo": ("csl_piezas", lambda ws: map_piezas(read_sheet(ws))),
        "Reportes": ("csl_reportes", lambda ws: map_reportes(read_sheet(ws))),
        "Inventario": ("csl_inventario", lambda ws: map_inventario(read_sheet(ws))),
        "Operadoras": ("csl_operadoras", lambda ws: map_operadoras(read_sheet(ws))),
        "LecturasSemanales": ("csl_lecturas_semanales", lambda ws: map_lecturas(read_sheet(ws))),
        "SesionesCliente": ("csl_sesiones_cliente", map_sesiones),
        "Credenciales": ("csl_credenciales", lambda ws: map_credenciales(read_sheet(ws))),
    }
    for sheet_name, (table, mapper) in sheet_map.items():
        if sheet_name in workbook.sheetnames:
            payloads[table] = mapper(workbook[sheet_name])
    if "SolicitudesEmpleo" in workbook.sheetnames:
        solicitudes, empleados = map_solicitudes(read_sheet(workbook["SolicitudesEmpleo"]))
        payloads["csl_solicitudes_empleo"] = solicitudes
        payloads["csl_empleados"] = empleados
    if "csl_lecturas_semanales" in payloads and "csl_sesiones_cliente" in payloads:
        payloads["csl_auditorias_semanales"] = build_auditorias(
            payloads["csl_lecturas_semanales"],
            payloads["csl_sesiones_cliente"],
        )
    return payloads


def request_json(url: str, service_key: str, method: str = "GET", payload: Any = None) -> Any:
    body = json.dumps(payload, ensure_ascii=False).encode("utf-8") if payload is not None else None
    request = urllib.request.Request(
        url,
        data=body,
        method=method,
        headers={
            "apikey": service_key,
            "Authorization": f"Bearer {service_key}",
            "Content-Type": "application/json",
            "Prefer": "resolution=merge-duplicates,return=minimal",
        },
    )
    try:
        with urllib.request.urlopen(request, timeout=60) as response:
            content = response.read().decode("utf-8")
            return json.loads(content) if content else None
    except urllib.error.HTTPError as error:
        details = error.read().decode("utf-8", errors="replace")
        raise RuntimeError(f"{method} {url} -> {error.code}: {details}") from error


def upsert_table(supabase_url: str, service_key: str, table: str, rows: list[dict[str, Any]], chunk_size: int) -> int:
    if not rows:
        return 0
    key = TABLE_KEYS[table]
    endpoint = f"{supabase_url.rstrip('/')}/rest/v1/{table}?on_conflict={urllib.parse.quote(key)}"
    total = 0
    for start in range(0, len(rows), chunk_size):
        chunk = rows[start : start + chunk_size]
        request_json(endpoint, service_key, method="POST", payload=chunk)
        total += len(chunk)
        time.sleep(0.05)
    return total


def count_table(supabase_url: str, service_key: str, table: str) -> int:
    endpoint = f"{supabase_url.rstrip('/')}/rest/v1/{table}?select=*&limit=1"
    request = urllib.request.Request(
        endpoint,
        method="GET",
        headers={
            "apikey": service_key,
            "Authorization": f"Bearer {service_key}",
            "Range-Unit": "items",
            "Range": "0-0",
            "Prefer": "count=exact",
        },
    )
    with urllib.request.urlopen(request, timeout=60) as response:
        content_range = response.headers.get("Content-Range", "0-0/0")
        return int(content_range.rsplit("/", 1)[-1])


def main() -> int:
    parser = argparse.ArgumentParser()
    parser.add_argument("--excel", default=str(DEFAULT_EXCEL_PATH))
    parser.add_argument("--dry-run", action="store_true")
    parser.add_argument("--only-pulsos", action="store_true")
    parser.add_argument("--chunk-size", type=int, default=500)
    args = parser.parse_args()

    excel_path = Path(args.excel)
    if not excel_path.exists():
        raise SystemExit(f"No existe el Excel: {excel_path}")

    payloads = build_payloads(excel_path)
    if args.only_pulsos:
        pulsos_tables = {"csl_operadoras", "csl_lecturas_semanales", "csl_sesiones_cliente", "csl_auditorias_semanales"}
        payloads = {table: rows for table, rows in payloads.items() if table in pulsos_tables}
    print("Resumen preparado:")
    for table, rows in payloads.items():
        print(f"- {table}: {len(rows)} registros")

    if args.dry_run:
        return 0

    env = load_env()
    supabase_url = env.get("NEXT_PUBLIC_SUPABASE_URL", "")
    service_key = env.get("SUPABASE_SERVICE_ROLE_KEY", "")
    if not supabase_url or not service_key:
        raise SystemExit("Faltan NEXT_PUBLIC_SUPABASE_URL o SUPABASE_SERVICE_ROLE_KEY en .env.local")

    print("\nCargando a Supabase...")
    for table, rows in payloads.items():
        imported = upsert_table(supabase_url, service_key, table, rows, args.chunk_size)
        print(f"- {table}: {imported} upsert")

    print("\nConteos en Supabase:")
    for table in payloads:
        print(f"- {table}: {count_table(supabase_url, service_key, table)}")

    return 0


if __name__ == "__main__":
    sys.exit(main())
